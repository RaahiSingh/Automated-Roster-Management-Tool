import glob
from flask import Blueprint, current_app, render_template, request, session, flash, redirect, url_for,send_file
import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime, timedelta
import calendar
import tempfile
from app.email_util import process_excel_and_send_email

lead_bp = Blueprint('lead', __name__)
pending_leaves = []


lead_bp = Blueprint('lead', __name__)
pending_leaves = []

@lead_bp.route('/lead', methods=['GET', 'POST'])
def edit_shift():
    message = ''

    if request.method == 'POST' and 'confirm' in request.form:
        bulk_data = session.get('bulk_shift_data', [])
        filepath = None

        for entry in bulk_data:
            try:
                filename = f"template_{entry['month']}_{entry['year']}.xlsx"
                filepath = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', filename))

                df_raw = pd.read_excel(filepath, header=None)
                df = pd.read_excel(filepath, skiprows=3, header=None)
                wb = load_workbook(filepath)
                sheet = wb.active

                for i, row in df.iterrows():
                    if str(row[1]).strip().upper() == entry['signum'].upper() and str(row[4]).strip().upper() == entry['name'].upper():
                        start_col = None
                        for j in range(6, len(row)):
                            if str(df_raw.iloc[2, j]).strip() == entry['date']:
                                start_col = j
                                break

                        if start_col is not None:
                            excel_row = i + 4
                            excel_col = start_col + 1

                            if entry['shift'] in ['OC', 'L']:
                                sheet.cell(row=excel_row, column=excel_col).value = entry['shift']
                            else:
                                applied_days = 0
                                current_col = start_col
                                while applied_days < 5 and current_col < len(row):
                                    day_name = str(df_raw.iloc[1, current_col]).strip().lower()
                                    if day_name not in ['sat', 'sun']:
                                        sheet.cell(row=excel_row, column=current_col + 1).value = entry['shift']
                                        applied_days += 1
                                    current_col += 1

                wb.save(filepath)

            except Exception as e:
                message += f"Error updating {entry['name']} ({entry['signum']}): {str(e)}\n"

        # Try sending email after saving the Excel file
        if filepath:
            email_status = process_excel_and_send_email(filepath)
            if email_status:
                message += "\nOn-Call Summary Email sent successfully."
            else:
                message += "\n⚠ On-Call Email sending failed or no OC data found."

        session.pop('bulk_shift_data', None)

        if not message:
            message = "All shifts successfully updated and email sent."

        return render_template('lead.html', message=message)

    elif request.method == 'POST':
        names = request.form.getlist('name')
        signums = request.form.getlist('signum')
        dates = request.form.getlist('date')
        shifts = request.form.getlist('shift')
        month = request.form.get('month')
        year = request.form.get('year')

        bulk_data = []
        for i in range(len(names)):
            if names[i] and signums[i] and dates[i] and shifts[i]:
                bulk_data.append({
                    'name': names[i].strip().upper(),
                    'signum': signums[i].strip().upper(),
                    'date': dates[i].strip(),
                    'shift': shifts[i].strip().upper(),
                    'month': month,
                    'year': year
                })

        session['bulk_shift_data'] = bulk_data
        flash(f"{len(bulk_data)} entries added. You can preview before saving.", "info")
        return redirect(url_for('lead.edit_shift'))

    return render_template('lead.html')



@lead_bp.route('/upload-template', methods=['GET', 'POST'])
def upload_team_template():
    team_file='template.xlsx'
    if request.method == 'POST':
        file = request.files.get(team_file)

        if file:
            df = pd.read_excel(file)

            # Make sure the column names in Excel are correct
            required_cols = {'Function', 'Signum', 'Location'}
            if not required_cols.issubset(set(df.columns)):
                flash("Excel must contain columns: Name, Signum, Location", "danger")
                return redirect(request.url)

            # Convert to list of dicts
            team_data = df[['Function', 'Signum', 'Location']].to_dict(orient='records')
            session['team_data'] = team_data

            flash("Team template uploaded successfully.", "success")
            return redirect(url_for('lead.create_roster'))  # route where the form is shown

    return render_template('upload_template.html')  # your upload page

@lead_bp.route('/create-roster', methods=['GET', 'POST'])
def create_roster():
    if 'team_data' not in session:
        flash("Please upload team data first.", "warning")
        return redirect(url_for('lead.upload_team_template'))

    # Default values (for GET)
    month = request.args.get('month', 'June')
    year = int(request.args.get('year', 2025))

    # Convert month name to number
    month_num = list(calendar.month_name).index(month)
    num_days = calendar.monthrange(year, month_num)[1]

    # Generate day headers (e.g., ['Mon', 'Tue', ...])
    days = []
    for day in range(1, num_days + 1):
        weekday = calendar.day_abbr[calendar.weekday(year, month_num, day)]
        days.append({'date': day, 'weekday': weekday})

    return render_template("create_roster.html", days=days, selected_month=month, selected_year=year)



@lead_bp.route('/preview')
def preview():
    try:
        if 'bulk_shift_data' not in session or not session['bulk_shift_data']:
            flash("No shift data found to preview.", "warning")
            return render_template('preview.html', data=[], full_data=None)

        entry = session['bulk_shift_data'][-1]
        filename = f"template_{entry['month']}_{entry['year']}.xlsx"
        filepath = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', filename))

        df = pd.read_excel(filepath, skiprows=2)
        df.fillna("", inplace=True)

        return render_template('preview.html', data=session['bulk_shift_data'], full_data=df.to_dict(orient='records'))

    except Exception as e:
        flash(f"Error reading Excel file: {e}", "danger")
        return render_template('preview.html', data=[], full_data=None)

@lead_bp.route('/lead/leaves', methods=['GET', 'POST'])
def view_pending_leaves():
    if request.method == 'POST':
        signum = request.form.get('signum')
        action = request.form.get('action')

        for req in pending_leaves:
            if req['signum'] == signum and req['status'] == 'Pending':
                req['status'] = 'Approved' if action == 'approve' else 'Rejected'
                break

        flash(f'Leave request for {signum} has been {action}d.', 'success')

    return render_template('manage_leaves.html', leaves=pending_leaves)

@lead_bp.route('/lead/leave-calendar')
def leave_calendar():
    date_map = {}

    for req in pending_leaves:
        start = datetime.strptime(req['start'], "%Y-%m-%d")
        end = datetime.strptime(req['end'], "%Y-%m-%d")

        delta = (end - start).days + 1
        for i in range(delta):
            day = start + timedelta(days=i)
            day_str = day.strftime("%Y-%m-%d")
            if day_str not in date_map:
                date_map[day_str] = []
            date_map[day_str].append(f"{req['name']}")

    return render_template('leave_calendar.html', calendar=date_map)

@lead_bp.route('/lead/analytics', methods=['GET', 'POST'], endpoint='lead_analytics')
def lead_analytics():
  

    filters = {
        'employee': request.form.get('employee') if request.method == 'POST' else '',
        'group': request.form.get('group') if request.method == 'POST' else '',
        'shift_type': request.form.get('shift_type') if request.method == 'POST' else '',
        'month': request.form.get('month') if request.method == 'POST' else '',
        'year': request.form.get('year') if request.method == 'POST' else ''
    }

    if filters['month'] and filters['year']:
        pattern = f"Shift_Plan_Sample_{filters['month']}_{filters['year']}*.xlsx"
    else:
        pattern = "Shift_Plan_*.xlsx"

    excel_files = glob.glob(os.path.join(current_app.root_path, 'static', 'Shift_Plan_*.xlsx'))
    if not excel_files:
        flash("Excel data not found.", "danger")
        return redirect(url_for('lead.edit_shift'))

    latest_file = max(excel_files, key=os.path.getmtime)

    try:
       
        raw_title = pd.read_excel(latest_file, header=None, nrows=1).iloc[0, 0]
        month = raw_title.split("-")[-1].strip() if isinstance(raw_title, str) else "Month Not Found"
        df = pd.read_excel(latest_file, skiprows=2)
        fixed_columns = ["function", "signum", "lc", "oc", "name", "location", "working_days"]
        dynamic_columns = [f"day_{i}" for i in range(df.shape[1] - len(fixed_columns))]

        df.columns = fixed_columns + dynamic_columns
        
        df = df[~df['function'].isin(['M', 'E1', 'E2', 'OC', 'L', 'G', 'N', 'H', ''])]
        df = df[~df['name'].isin(['nan', ''])] 
        if filters['month'] and filters['year']:
            capitalized_month = filters['month'].capitalize()
            if capitalized_month not in calendar.month_name:
                flash(f"Invalid month: {capitalized_month}", "danger")
                return redirect(url_for('lead.edit_shift'))
        
            month_num = list(calendar.month_name).index(capitalized_month)
            year = int(filters['year'])
        else:
            now = datetime.now()
            month_num = now.month
            year = now.year       

        num_days = calendar.monthrange(year, month_num)[1]
        pretty_day_columns = [
            f"{day} ({calendar.day_abbr[calendar.weekday(year, month_num, day)]})" for day in range(1, num_days + 1)
        ]
        rename_map = dict(zip(dynamic_columns, pretty_day_columns))
        df.rename(columns=rename_map, inplace=True)
        df.fillna("", inplace=True)

        df['name'] = df['name'].astype(str).str.strip()
        df = df[df['name'] != ""]


        if filters['employee']:
            df = df[df['name'].str.contains(filters['employee'], case=False, na=False)]
        if filters['group']:
            df = df[df['function'].str.contains(filters['group'], case=False, na=False)]
        if filters['shift_type']:
            df = df[df[list(rename_map.values())].apply(
                lambda row: row.astype(str).str.contains(filters['shift_type'], case=False, na=False).any(), axis=1
            )]

        df[list(rename_map.values())] = df[list(rename_map.values())].astype(str)
        shift_counts = pd.Series(df[list(rename_map.values())].values.ravel()).value_counts().to_dict()

        analytics_data = {
            'summary': {
                'month': filters['month'].capitalize() if filters['month'] else "Unknown",
                'year': filters['year'] or datetime.now().year,
                'total_employees': df['name'].nunique(),
                'total_leaves': shift_counts.get('L', 0),
                'general_shifts': shift_counts.get('G', 0),
                'night_shifts': shift_counts.get('N', 0),
                'e1_shifts': shift_counts.get('E1', 0),
                'e2_shifts': shift_counts.get('E2', 0)
            }
        }
        table_columns = ["function", "signum", "name", "location"] + list(rename_map.values())
        


        table_rows = df[table_columns].to_dict(orient='records')

        return render_template(
            'lead_analytics.html',
            filters=filters,
            data=analytics_data,
            month=month,
            table_rows=table_rows,
            columns=table_columns
        )

    except Exception as e:
        flash(f"Error processing Excel: {e}", "danger")
        return redirect(url_for('lead.edit_shift'))



@lead_bp.route('/lead/download-summary-excel', methods=['POST'])
def download_summary_excel():
    filters = {
        'employee': request.form.get('employee', ''),
        'group': request.form.get('group', ''),
        'shift_type': request.form.get('shift_type', ''),
        'month': request.form.get('month', ''),
        'year': request.form.get('year', '')
    }

    try:
        pattern = f"Shift_Plan_Sample_{filters['month']}_{filters['year']}*.xlsx" if filters['month'] and filters['year'] else "Shift_Plan_*.xlsx"
        excel_files = glob.glob(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', pattern)))
        if not excel_files:
            flash("No file found for this month/year.", "danger")
            return redirect(url_for('lead.lead_analytics'))

        latest_file = max(excel_files, key=os.path.getmtime)

        df = pd.read_excel(latest_file, skiprows=2)
        fixed_columns = ["function", "signum", "lc", "oc", "name", "location", "working_days"]
        dynamic_columns = [f"day_{i}" for i in range(df.shape[1] - len(fixed_columns))]
        df.columns = fixed_columns + dynamic_columns

        # Filtering
        df = df[~df['function'].isin(['M', 'E1', 'E2', 'OC', 'L', 'G', 'N', 'H', ''])]
        df['name'] = df['name'].astype(str).str.strip()
        df = df[df['name'] != ""]

        if filters['employee']:
            df = df[df['name'].str.contains(filters['employee'], case=False, na=False)]
        if filters['group']:
            df = df[df['function'].str.contains(filters['group'], case=False, na=False)]
        if filters['shift_type']:
            df[dynamic_columns] = df[dynamic_columns].astype(str)
            df = df[df[dynamic_columns].apply(
                lambda row: row.str.contains(filters['shift_type'], case=False, na=False).any(), axis=1
            )]

        # Create day column names like "1 (Sun)"
        if filters['month'] and filters['year']:
            month_num = list(calendar.month_name).index(filters['month'].capitalize())
            year = int(filters['year'])
        else:
            now = datetime.now()
            month_num = now.month
            year = now.year

        num_days = calendar.monthrange(year, month_num)[1]
        pretty_day_columns = [
            f"{day} ({calendar.day_abbr[calendar.weekday(year, month_num, day)]})"
            for day in range(1, num_days + 1)
        ]
        rename_map = dict(zip(dynamic_columns[:len(pretty_day_columns)], pretty_day_columns))

        df.rename(columns=rename_map, inplace=True)
        day_columns = list(rename_map.values())

        # Summary
        shift_counts = pd.Series(df[day_columns].values.ravel()).value_counts().to_dict()
        summary_df = pd.DataFrame([{
            'Month': filters['month'].capitalize() if filters['month'] else "Unknown",
            'Year': filters['year'] or datetime.now().year,
            'Total Employees': df['name'].nunique(),
            'Leaves (L)': shift_counts.get('L', 0),
            'General (G)': shift_counts.get('G', 0),
            'Night (N)': shift_counts.get('N', 0),
            'E1': shift_counts.get('E1', 0),
            'E2': shift_counts.get('E2', 0)
        }])

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                summary_df.to_excel(writer, index=False, sheet_name='Summary')
                df_to_save = df[["function", "signum", "name", "location"] + day_columns]
                df_to_save.to_excel(writer, index=False, sheet_name='Filtered Roster')

            tmp.seek(0)
            return send_file(tmp.name, as_attachment=True, download_name="Shift_Summary_With_Roster.xlsx")

    except Exception as e:
        flash(f"Error generating summary: {e}", "danger")
        return redirect(url_for('lead.lead_analytics'))
