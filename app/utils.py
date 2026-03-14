from openpyxl import load_workbook

def fill_excel_template(template_path, output_path, members):
    wb = load_workbook(template_path)
    ws = wb.active
    for i, member in enumerate(members, start=4):
        ws.cell(row=i, column=5).value = member['name']
        ws.cell(row=i, column=2).value = member['signum']
        ws.cell(row=i, column=6).value = member['location']
        ws.cell(row=i, column=1).value = member['function']
    wb.save(output_path)
